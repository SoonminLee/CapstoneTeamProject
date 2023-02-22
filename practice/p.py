import re
import openpyxl

# str = "1회제공량 5개(29g) 총약3.5회 제공량(104g) 1회 제공량당 함량 *%영양소 기준치: 열량 151kcal,탄수화물20g 6%,당류8g, 단백질 2g 4%,지방7g 14%, 포화지방4.46g 29%, 트랜스지방 0.5g미만, 콜레스테롤5mg미만 2%, 나트륨 90mg 5%, *%영양소 기준치:1일 영양소 기준치에대한 비율"
wb = openpyxl.load_workbook('Nutrition_Facts1.xlsx')
ws = wb.active

pattern1 = "(열량 \d*\.?\d+)"
a = re.search(pattern1, ws['C2'].value).group(0)
a = re.search('\d*\.?\d+',a).group(0)
pattern2 = "(탄수화물\d*\.?\d+)"
b = re.search(pattern2, ws['C2'].value).group(0)
b = re.search('\d*\.?\d+',b).group(0)
pattern3 = "(당류\d*\.?\d+)"
c = re.search(pattern3, ws['C2'].value).group(0)
c = re.search('\d*\.?\d+',c).group(0)
pattern4 = "(단백질 \d*\.?\d+)"
d = re.search(pattern4, ws['C2'].value).group(0)
d = re.search('\d*\.?\d+',d).group(0)
pattern5 = "(지방\d*\.?\d+)"
e = re.search(pattern5, ws['C2'].value).group(0)
e = re.search('\d*\.?\d+',e).group(0)
pattern6 = "(포화지방\d*\.?\d+)"
f = re.search(pattern6, ws['C2'].value).group(0)
f = re.search('\d*\.?\d+',f).group(0)
pattern7 = "(트랜스지방 \d*\.?\d+)"
g = re.search(pattern7, ws['C2'].value).group(0)
g = re.search('\d*\.?\d+',g).group(0)
pattern8 = "(콜레스테롤\d*\.?\d+)"
h = re.search(pattern8, ws['C2'].value).group(0)
h = re.search('\d*\.?\d+',h).group(0)
pattern9 = "(나트륨 \d*\.?\d+)"
i = re.search(pattern9, ws['C2'].value).group(0)
i = re.search('\d*\.?\d+',i).group(0)
wb = openpyxl.Workbook()
# ws = wb.create_sheet('hello')
ws = wb.active
ws.title = 'hello'
ws['A1'] = '칼로리[kcal]'
ws['B1'] = '탄수화물[g]'
ws['C1'] = '당류[g]'
ws['D1'] = '단백질[g]'
ws['E1'] = '지방[g]'
ws['F1'] = '포화지방[g]'
ws['G1'] = '트랜스지방[g]'
ws['H1'] = '콜레스테롤[mg]'
ws['I1'] = '나트륨[mg]'
ws['A2'] = a
ws['B2'] = b
ws['C2'] = c
ws['D2'] = d
ws['E2'] = e
ws['F2'] = f
ws['G2'] = g
ws['H2'] = h
ws['I2'] = i
wb.save('testttt.xlsx')

# print(a)
# print(b)
# print(c)
# print(d)
# print(e)
# print(f)
# print(g)
# print(h)
# print(i)
