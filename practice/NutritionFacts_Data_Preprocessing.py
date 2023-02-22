import openpyxl
import cv2
import pyzbar.pyzbar as pyzbar
from playsound import playsound
import pandas as pd
import requests
from bs4 import BeautifulSoup
import re

# 바코드를 영양성분표로
def barcode_to_NutritionFacts():
  # 카메라로 바코드를 추출
  a=0
  used_codes = []
  data_list = []

  try:
      f = open("qr_barcode_data.txt", "r", encoding="utf8")
      data_list = f.readlines()
  except FileNotFoundError: 
      pass
  else:
      f.close()

  cap = cv2.VideoCapture(0)
  cap.set(3, 640)
  cap.set(4, 480)

  for i in data_list:
      used_codes.append(i.rstrip('\n'))

  while a==0:
      success, frame = cap.read()
      for code in pyzbar.decode(frame):
          cv2.imwrite('qr_barcode_image.png', frame)
          my_code = code.data.decode('utf-8')
          if my_code not in used_codes:
            print("인식 성공 : ", my_code)
            playsound("qrbarcode_beep.mp3")
            used_codes.append(my_code)
            
            f2 = open("qr_barcode_data.txt", "w", encoding="utf8")
            f2.write(my_code+'\n')
            f2.close()
            a+=1
          elif my_code in used_codes:
            print("이미 인식된 코드 입니다")
            playsound('qrbarcode_beep.mp3')
          else:
              pass
      # esc키 눌렀을시 종료
      key = cv2.waitKey(1)    
      if key == 27:
          break
      cv2.imshow('QRcode Barcode Scan', frame)
      cv2.waitKey(1)

  # 바코드를 입력 받아 품목보고번호를 검색
  BAR_CD = my_code
  url = 'http://openapi.foodsafetykorea.go.kr/api/14adc49a10ff4df2a831/I2570/xml/1/5/BRCD_NO=' + BAR_CD
  def parse():
    try:
        PRDLST_REPORT_NO = item.find("PRDLST_REPORT_NO").get_text()
        
        return {
            PRDLST_REPORT_NO
        }
    except AttributeError as e:
        return {
            None
        }
  result = requests.get(url)
  soup = BeautifulSoup(result.content, 'lxml-xml')
  items = soup.find_all("I2570")
# 파싱한 내용에서 품목보고번호만 추출
  row = []
  for item in items:
    row.append(parse())
  string = str(row)
  Produtnumbers = re.sub(r'[^0-9]', '', string)
  # print(Produtnumbers)

# 품목보고번호를 입력 받아 영양성분을 검색 후 제품명과 영양성분 csv파일로 출력
  url = 'http://apis.data.go.kr/B553748/CertImgListService/getCertImgListService'
  params ={
'serviceKey' : 'IyQg8I2dXbv8kkUs2Gki35cm64Cu+xaUWkNCsFipH3WWV6/iZD4HHrq4v+ykezvft92l9H5S0zULIYrQonfaUA==',
'prdlstReportNo' : Produtnumbers
 }

  def parse():
        try:
          PRODLSTNM = item.find("prdlstNm").get_text()
          nutrient = item.find("nutrient").get_text()
          return {
            "제품명":PRODLSTNM,
            "영양성분":nutrient,
        }
        except AttributeError as e:
          return {
            "제품명":None,
            "영양성분":None
        }

  result = requests.get(url, params=params)
  soup = BeautifulSoup(result.content, 'lxml-xml')
  items = soup.find_all("item")
  print(items)

  row = []
  for item in items:
      row.append(parse())
  df = pd.DataFrame(row)
  df.to_excel("Before_Nutrition_Facts.xlsx")
  print(row)
def NutritionFacts_Data_Preprocessing():
  wb = openpyxl.load_workbook('Before_Nutrition_Facts.xlsx')
  ws = wb.active
  pattern1 = "(열량 \d*\.?\d+)"
  a = re.search(pattern1, ws['C2'].value).group(0)
  a = re.search('\d*\.?\d+',a).group(0)
  pattern2 = "(탄수화물\d*\.?\d+)"
  b = re.search(pattern2, ws['C2'].value).group(0)
  b = re.search('\d*\.?\d+',b).group(0)
  pattern3 = "(당류\d*\.?\d+)"
  c = re.search(pattern3, ws['C2'].value).group(0)
  c = re.search('\d*\.?\d+',c).group(0)
  pattern4 = "(단백질 \d*\.?\d+)"
  d = re.search(pattern4, ws['C2'].value).group(0)
  d = re.search('\d*\.?\d+',d).group(0)
  pattern5 = "(지방\d*\.?\d+)"
  e = re.search(pattern5, ws['C2'].value).group(0)
  e = re.search('\d*\.?\d+',e).group(0)
  pattern6 = "(포화지방\d*\.?\d+)"
  f = re.search(pattern6, ws['C2'].value).group(0)
  f = re.search('\d*\.?\d+',f).group(0)
  pattern7 = "(트랜스지방 \d*\.?\d+)"
  g = re.search(pattern7, ws['C2'].value).group(0)
  g = re.search('\d*\.?\d+',g).group(0)
  pattern8 = "(콜레스테롤\d*\.?\d+)"
  h = re.search(pattern8, ws['C2'].value).group(0)
  h = re.search('\d*\.?\d+',h).group(0)
  pattern9 = "(나트륨 \d*\.?\d+)"
  i = re.search(pattern9, ws['C2'].value).group(0)
  i = re.search('\d*\.?\d+',i).group(0)
  wb = openpyxl.Workbook()
  ws = wb.active
  ws.title = 'barcode_NutritionFacts'
  ws['A1'] = '칼로리[kcal]'
  ws['B1'] = '탄수화물[g]'
  ws['C1'] = '당류[g]'
  ws['D1'] = '단백질[g]'
  ws['E1'] = '지방[g]'
  ws['F1'] = '포화지방[g]'
  ws['G1'] = '트랜스지방[g]'
  ws['H1'] = '콜레스테롤[mg]'
  ws['I1'] = '나트륨[mg]'
  ws['A2'] = a
  ws['B2'] = b
  ws['C2'] = c
  ws['D2'] = d
  ws['E2'] = e
  ws['F2'] = f
  ws['G2'] = g
  ws['H2'] = h
  ws['I2'] = i
  wb.save('NutritionFacts.xlsx')

barcode_to_NutritionFacts()
NutritionFacts_Data_Preprocessing()
